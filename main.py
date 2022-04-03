from datetime import datetime
import os
import re

import openpyxl
import twstock


if __name__ == "__main__":
    twstock.realtime.mock = False
    file_name = '股票時間成交量找出起漲股.xlsx'
    if not os.path.exists(file_name):
        input(f'找不到檔案 - {file_name} ，請確認是否與程式放在同一資料夾底下... 按 Enter 退出')
        exit()
    try:
        workbook = openpyxl.load_workbook(file_name)
        print('資料抓取中')
        for sheet in workbook:
            if not re.search(r'\(\d+\)', sheet.title):
                continue
            stock_dict = {}
            for i in range(sheet.max_column // 8):
                for j in range(sheet.max_row // 20):
                    col = i * 9 + 3
                    row = j * 20 + 1
                    cell = sheet.cell(row=row, column=col)
                    stock_code = cell.value
                    if stock_code:
                        stock_dict[str(stock_code)] = (col, row)
            result = twstock.realtime.get(list(stock_dict.keys()))
            for code, index in stock_dict.items():
                col, row = index
                if code not in result or not result[code]['success'] or code == 'success':
                    continue
                info_datetime = datetime.fromtimestamp(result[code]['timestamp'])
                if info_datetime >= info_datetime.replace(hour=13, minute=45, second=0, microsecond=0):
                    info_datetime = info_datetime.replace(hour=13, minute=30, second=0, microsecond=0)
                session_num = int((info_datetime - info_datetime.replace(hour=9, minute=0, second=0, microsecond=0)).total_seconds() // 900) + 1
                sheet.cell(row=(row + session_num), column=col).value = result[code]['realtime']['accumulate_trade_volume']

    except Exception as e:
        print(f'Error: {e}')

    finally:

        workbook.save(file_name)
        workbook.close()
        print('抓取完成')